from fastapi import FastAPI, Query, HTTPException
import requests, io, json
from openpyxl import load_workbook

app = FastAPI()

@app.get("/fetch_excel_from_github")
def fetch_excel_from_github(path: str = Query(...)):
    try:
        # 1️⃣ GitHub RAW URL
        base_raw_url = "https://raw.githubusercontent.com/tousuienfurukawa-web/tousuien-hub/main/"
        url = base_raw_url + path.replace(" ", "%20")

        res = requests.get(url, timeout=10)
        res.raise_for_status()

        # 2️⃣ Excelを読み取り専用で開く（Vercelで安定）
        try:
            workbook = load_workbook(io.BytesIO(res.content), data_only=True, read_only=True)
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Excel読み込みエラー: {e}")

        sheet = workbook.active
        headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        records = [dict(zip(headers, row)) for row in sheet.iter_rows(min_row=2, values_only=True)]

        # 3️⃣ データ整形（最初の10件のみ）
        summary = [
            {
                "企業コード": r.get("企業コード"),
                "会社名": r.get("会社名"),
                "合計金額": f"{r.get('通貨')} {r.get('合計金額')}" if r.get("通貨") else r.get("合計金額"),
                "入金確認日": r.get("入金確認日"),
                "Invoice番号": r.get("Invoice番号")
            }
            for r in records if r.get("企業コード")
        ]

        return summary[:10]

    except requests.exceptions.RequestException as e:
        raise HTTPException(status_code=400, detail=f"GitHubファイル取得エラー: {e}")

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"サーバー内部エラー: {e}")
