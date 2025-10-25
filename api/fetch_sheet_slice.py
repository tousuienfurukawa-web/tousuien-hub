# api/fetch_sheet_slice.py
"""
Fetch a slice of an Excel sheet hosted on GitHub raw and return JSON.
Supports:
- query params: sheet, start_row, end_row, cols (A:C or header names), summary (true/false), limit
- caching: Redis via cache_util (if available) and in-process memory fallback
- safe conversion of Excel cell values to JSON serializable forms
"""
from http.server import BaseHTTPRequestHandler
from urllib.parse import urlparse, parse_qs
import json
import io
import os
import re
import hashlib
import time
import threading

import openpyxl
import requests
from datetime import datetime, date

# Optional cache util module (expects cache_util.cache_get / cache_set)
try:
    import cache_util
except Exception:
    cache_util = None

# Configuration
GITHUB_XLSX_URL = os.environ.get(
    "GITHUB_XLSX_URL",
    "https://raw.githubusercontent.com/tousuienfurukawa-web/tousuien-hub/main/data/Customer_Management_latest.xlsx"
)
CACHE_TTL = int(os.environ.get("CACHE_TTL_SECONDS", "300"))
MAX_MEM_CACHE_ENTRIES = int(os.environ.get("MAX_MEM_CACHE_ENTRIES", "200"))

# ---------------------------
# In-process memory cache (fallback)
# ---------------------------
MEM_CACHE = {}   # key -> { 'ts': epoch, 'val': obj_or_jsonstr, 'is_json_str': bool }
MEM_LOCK = threading.Lock()

def mem_cache_get(key):
    with MEM_LOCK:
        entry = MEM_CACHE.get(key)
        if not entry:
            return None
        if (time.time() - entry['ts']) <= CACHE_TTL:
            val = entry['val']
            if entry.get('is_json_str'):
                try:
                    return json.loads(val)
                except Exception:
                    return val
            return val
        # expired
        try:
            del MEM_CACHE[key]
        except KeyError:
            pass
        return None

def mem_cache_set(key, obj):
    is_json_str = False
    store_val = obj
    try:
        json.dumps(obj, ensure_ascii=False)
    except Exception:
        try:
            store_val = json.dumps(obj, default=str, ensure_ascii=False)
            is_json_str = True
        except Exception:
            store_val = str(obj)
            is_json_str = True

    with MEM_LOCK:
        MEM_CACHE[key] = {'ts': time.time(), 'val': store_val, 'is_json_str': is_json_str}
        # Evict oldest if too many entries
        if len(MEM_CACHE) > MAX_MEM_CACHE_ENTRIES:
            oldest_key = min(MEM_CACHE.items(), key=lambda kv: kv[1]['ts'])[0]
            try:
                del MEM_CACHE[oldest_key]
            except KeyError:
                pass

# ---------------------------
# Utilities
# ---------------------------
def normalize_header(s):
    if s is None:
        return ""
    return str(s).replace("\n", " ").replace("　", " ").strip()

def col_letter_to_index(col):
    col = col.upper()
    idx = 0
    for ch in col:
        if 'A' <= ch <= 'Z':
            idx = idx * 26 + (ord(ch) - ord('A') + 1)
    return idx - 1

def parse_col_range(cols_spec, headers):
    if not cols_spec:
        return list(range(len(headers)))
    cols_spec = cols_spec.strip()
    if ':' in cols_spec and re.match(r'^[A-Za-z]+:[A-Za-z]+$', cols_spec):
        a, b = cols_spec.split(':')
        start = col_letter_to_index(a)
        end = col_letter_to_index(b)
        if start < 0:
            start = 0
        if end >= len(headers):
            end = len(headers) - 1
        return list(range(start, end + 1))
    names = [x.strip() for x in cols_spec.split(',') if x.strip()]
    indices = []
    for name in names:
        name_norm = normalize_header(name).lower()
        found = False
        for i, h in enumerate(headers):
            if normalize_header(h).lower() == name_norm:
                indices.append(i)
                found = True
                break
        if not found:
            for i, h in enumerate(headers):
                if name_norm in normalize_header(h).lower():
                    indices.append(i)
                    found = True
                    break
    return indices

def convert_value(value):
    """Convert Excel cell value into JSON-serializable Python value."""
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        try:
            return value.isoformat()
        except Exception:
            return str(value)
    if isinstance(value, (int, float, str, bool)):
        return value
    # for bytes etc.
    try:
        return str(value)
    except Exception:
        return ""

def make_cache_key(sheet, start_row, end_row, cols_spec, summary, limit):
    raw = f"sheet={sheet}|start={start_row}|end={end_row}|cols={cols_spec or ''}|summary={summary}|limit={limit}"
    h = hashlib.sha256(raw.encode('utf-8')).hexdigest()
    return f"sheet_slice:{h}"

# ---------------------------
# HTTP Handler
# ---------------------------
class handler(BaseHTTPRequestHandler):
    def do_GET(self):
        try:
            parsed = urlparse(self.path)
            params = parse_qs(parsed.query)
            sheet = params.get('sheet', [None])[0]
            start_row = int(params.get('start_row', ['2'])[0])
            end_row = int(params.get('end_row', ['200'])[0])
            cols_spec = params.get('cols', [None])[0]
            summary = params.get('summary', ['false'])[0].lower() == 'true'
            limit = int(params.get('limit', ['10'])[0])

            if not sheet:
                self.send_response(400)
                self.send_header('Content-Type', 'application/json; charset=utf-8')
                self.end_headers()
                self.wfile.write(json.dumps({"error": "missing sheet param"}).encode('utf-8'))
                return

            # ---- Cache lookup: try Redis first, then memcache ----
            cache_key = make_cache_key(sheet, start_row, end_row, cols_spec, summary, limit)
            cached = None
            cache_source = None
            if cache_util is not None:
                try:
                    cached = cache_util.cache_get(cache_key)
                    if cached is not None:
                        cache_source = 'redis'
                except Exception:
                    cached = None

            if cached is None:
                try:
                    cached = mem_cache_get(cache_key)
                    if cached is not None:
                        cache_source = 'mem'
                except Exception:
                    cached = None

            if cached:
                if isinstance(cached, dict):
                    cached_meta = cached.get("_meta", {})
                    cached_meta["cached"] = True
                    cached["_meta"] = cached_meta
                self.send_response(200)
                self.send_header('Content-Type', 'application/json; charset=utf-8')
                self.send_header('Access-Control-Allow-Origin', '*')
                if cache_source == 'redis':
                    self.send_header('X-Cache', 'HIT-REDIS')
                else:
                    self.send_header('X-Cache', 'HIT-MEM')
                self.end_headers()
                self.wfile.write(json.dumps(cached, ensure_ascii=False).encode('utf-8'))
                return

            # ---- Not cached: fetch Excel and compute ----
            resp = requests.get(GITHUB_XLSX_URL, timeout=30)
            resp.raise_for_status()
            wb = openpyxl.load_workbook(io.BytesIO(resp.content), data_only=True, read_only=True)

            if sheet not in wb.sheetnames:
                self.send_response(404)
                self.send_header('Content-Type', 'application/json; charset=utf-8')
                self.end_headers()
                self.wfile.write(json.dumps({"error": "sheet not found"}).encode('utf-8'))
                return

            ws = wb[sheet]
            first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
            headers = [normalize_header(h) for h in (first_row or [])]
            col_indices = parse_col_range(cols_spec, headers)
            if not col_indices:
                col_indices = list(range(len(headers))) if headers else []

            rows = []
            total_count = 0
            end_row_cap = min(end_row, start_row + 20000)

            for i, row in enumerate(ws.iter_rows(min_row=start_row, max_row=end_row_cap, values_only=True), start=start_row):
                rowvals = []
                if not row:
                    for ci in col_indices:
                        rowvals.append("")
                else:
                    for ci in col_indices:
                        if ci < len(row):
                            rowvals.append(convert_value(row[ci]))
                        else:
                            rowvals.append("")
                rows.append(rowvals)
                total_count += 1
                if summary and total_count >= limit:
                    break

            result = {
                "sheet": sheet,
                "requested_rows": f"{start_row}-{end_row}",
                "returned": len(rows),
                "generated_at": datetime.utcnow().isoformat() + "Z",
                "_meta": {
                    "cached": False,
                    "cache_ttl": CACHE_TTL
                }
            }
            if summary:
                result["headers"] = [headers[i] for i in col_indices] if headers else []
                result["top_rows"] = rows[:limit]
                result["note"] = "summary mode - top N rows"
            else:
                result["headers"] = [headers[i] for i in col_indices] if headers else []
                result["rows"] = rows

            # ---- store in cache (Redis if available, always memcache) ----
            if cache_util is not None:
                try:
                    cache_util.cache_set(cache_key, result, ttl_seconds=CACHE_TTL)
                except Exception:
                    pass
            try:
                mem_cache_set(cache_key, result)
            except Exception:
                pass

            # return result
            self.send_response(200)
            self.send_header('Content-Type', 'application/json; charset=utf-8')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.send_header('X-Cache', 'MISS')
            self.end_headers()
            self.wfile.write(json.dumps(result, ensure_ascii=False).encode('utf-8'))

        except Exception as e:
            import traceback
            self.send_response(500)
            self.send_header('Content-Type', 'application/json; charset=utf-8')
            self.end_headers()
            error_response = {"error": str(e), "traceback": traceback.format_exc()}
            self.wfile.write(json.dumps(error_response, ensure_ascii=False).encode('utf-8'))
