# api/fetch_sales_summary_by_code.py
# read_only + short-term cache 版に「ヘッダー検出強化＋フォールバック」を追加した実装
from http.server import BaseHTTPRequestHandler
from urllib.parse import urlparse, parse_qs
import json
from datetime import datetime, date
import time
import threading
import os
import unicodedata

def convert_value(value):
    """セルの値を安全な形式に変換"""
    if value is None:
        return ""
    elif isinstance(value, (datetime, date)):
        return value.isoformat()
    elif isinstance(value, (int, float, str, bool)):
        return value
    else:
        return str(value)

# --- 短期キャッシュ（プロセス内） ---
CACHE = {}  # code -> {'ts': epoch_seconds, 'result': json_string}
CACHE_LOCK = threading.Lock()
CACHE_TTL = int(os.environ.get("CACHE_TTL", "300"))  # default 300s

def get_cached(code):
    with CACHE_LOCK:
        entry = CACHE.get(code)
        if not entry:
            return None
        now = time.time()
        if now - entry['ts'] <= CACHE_TTL:
            return entry['result']
        return None

def set_cache(code, json_str):
    with CACHE_LOCK:
        CACHE[code] = {'ts': time.time(), 'result': json_str}

def get_stale_if_any(code):
    with CACHE_LOCK:
        entry = CACHE.get(code)
        return entry['result'] if entry else None

# --- ヘッダー検出ユーティリティ ---
def normalize_header_text(s):
    """ヘッダー文字列を正規化して比較可能にする
    - 改行・タブをスペースに変換
    - 全角スペースを半角に
    - Unicode 正規化 (NFKC)
    - 連続空白除去、trim、小文字化
    """
    if s is None:
        return ""
    t = str(s)
    t = t.replace('\r', ' ').replace('\n', ' ').replace('\t', ' ')
    t = t.replace('　', ' ')
    t = unicodedata.normalize('NFKC', t)
    t = " ".join(t.split())
    return t.strip().lower()

def looks_like_company_code_header(h):
    """ヘッダー候補が企業コードに相当するかをあいまい判定する"""
    nh = normalize_header_text(h)
    if not nh:
        return False
    # 日本語パターン
    if '企業' in nh and 'コード' in nh:
        return True
    if '会社' in nh and 'コード' in nh:
        return True
    # 原料登録の特殊ケース
    if '原料' in nh and 'コード' in nh:
        return True
    if '原料資材' in nh and 'コード' in nh:
        return True
    # 英語表記候補
    if 'company' in nh and 'code' in nh:
        return True
    # 単純に 'コード' を含む場合も候補とする（慎重に扱う）
    if 'コード' in nh and len(nh.split()) <= 3:
        return True
    return False

class handler(BaseHTTPRequestHandler):
    def do_GET(self):
        try:
            parsed_path = urlparse(self.path)
            params = parse_qs(parsed_path.query)

            if 'code' not in params:
                self.send_response(400)
                self.send_header('Content-Type', 'application/json; charset=utf-8')
                self.end_headers()
                self.wfile.write(json.dumps({
                    "error": "Missing 'code' parameter",
                    "usage": "/api/fetch_sales_summary_by_code?code=BPR"
                }).encode('utf-8'))
                return

            code = params['code'][0].upper().strip()

            # 1) キャッシュ参照
            cached = get_cached(code)
            if cached:
                self.send_response(200)
                self.send_header('Content-Type', 'application/json; charset=utf-8')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.end_headers()
                try:
                    parsed = json.loads(cached)
                    parsed['_meta'] = parsed.get('_meta', {})
                    parsed['_meta']['cached'] = True
                    parsed['_meta']['cache_ttl'] = CACHE_TTL
                    out = json.dumps(parsed, ensure_ascii=False, indent=2)
                    self.wfile.write(out.encode('utf-8'))
                    return
                except Exception:
                    self.wfile.write(cached.encode('utf-8'))
                    return

            # 2) GitHub から Excel を取得
            import requests
            import io
            import openpyxl

            github_url = "https://raw.githubusercontent.com/tousuienfurukawa-web/tousuien-hub/main/data/Customer_Management_latest.xlsx"
            try:
                response = requests.get(github_url, timeout=30)
                response.raise_for_status()
            except Exception as e:
                stale = get_stale_if_any(code)
                if stale:
                    try:
                        parsed = json.loads(stale)
                    except Exception:
                        parsed = {"error": "stale cache exists but could not parse", "raw": stale}
                    parsed['_meta'] = parsed.get('_meta', {})
                    parsed['_meta']['cached'] = True
                    parsed['_meta']['cache_ttl'] = CACHE_TTL
                    parsed['_meta']['stale'] = True
                    parsed['_meta']['notice'] = f"GitHub fetch failed: {str(e)}"
                    self.send_response(200)
                    self.send_header('Content-Type', 'application/json; charset=utf-8')
                    self.send_header('Access-Control-Allow-Origin', '*')
                    self.end_headers()
                    out = json.dumps(parsed, ensure_ascii=False, indent=2)
                    self.wfile.write(out.encode('utf-8'))
                    return
                else:
                    self.send_response(500)
                    self.send_header('Content-Type', 'application/json; charset=utf-8')
                    self.end_headers()
                    error_response = json.dumps({
                        "error": "Failed to fetch Excel from GitHub and no cache available",
                        "detail": str(e),
                        "type": type(e).__name__
                    }, ensure_ascii=False)
                    self.wfile.write(error_response.encode('utf-8'))
                    return

            workbook = openpyxl.load_workbook(io.BytesIO(response.content), data_only=True, read_only=True)

            target_sheets = ["会社情報登録", "原料登録", "商品登録", "受注登録"]
            result = {
                "code": code,
                "found": False,
                "data": {},
                "_meta": {
                    "generated_at": datetime.utcnow().isoformat() + "Z",
                    "cached": False,
                    "cache_ttl": CACHE_TTL,
                    "detection": {}  # per-sheet detection info
                }
            }

            # 各シートを処理
            for sheet_name in target_sheets:
                detection_info = {"method": None, "header_row": None, "code_col_index": None}
                if sheet_name not in workbook.sheetnames:
                    result["data"][sheet_name] = {
                        "error": "シートが見つかりません",
                        "headers": [],
                        "rows": [],
                        "count": 0
                    }
                    result["_meta"]["detection"][sheet_name] = {"method": "sheet_not_found"}
                    continue

                ws = workbook[sheet_name]

                # 1) 上位数行（1〜5行）を見てヘッダー行を探す
                code_col_index = None
                headers = []
                header_row_index = None

                for r in range(1, 6):
                    try:
                        row_iter = ws.iter_rows(min_row=r, max_row=r, values_only=True)
                        row = next(row_iter, None)
                    except Exception:
                        row = None
                    if not row:
                        continue
                    cand_headers = [convert_value(cell) for cell in row]
                    non_empty = sum(1 for c in cand_headers if str(c).strip() != "")
                    if non_empty < 2:
                        # ヘッダーらしくない行ならスキップ
                        continue
                    # 各列を判定
                    for idx, h in enumerate(cand_headers):
                        if looks_like_company_code_header(h):
                            code_col_index = idx
                            headers = cand_headers
                            header_row_index = r
                            break
                    if code_col_index is not None:
                        break

                # 2) ヘッダーロジックで見つからなければ、まず1行目をヘッダー候補とする（既存処理互換）
                if header_row_index is None:
                    try:
                        first_row_iter = ws.iter_rows(min_row=1, max_row=1, values_only=True)
                        first_row = next(first_row_iter, None)
                    except Exception:
                        first_row = None
                    if first_row and first_row[0] is not None:
                        headers = [convert_value(cell) for cell in first_row]
                        # try to locate company code header in first row (fallback)
                        for idx, h in enumerate(headers):
                            if looks_like_company_code_header(h):
                                code_col_index = idx
                                header_row_index = 1
                                break
                    else:
                        headers = []

                matched_rows = []

                # 3) もし code_col_index が見つかれば、ヘッダー行に続く行から抽出
                if code_col_index is not None:
                    detection_info['method'] = 'header_match'
                    detection_info['header_row'] = header_row_index
                    detection_info['code_col_index'] = code_col_index
                    # 開始行：ヘッダー行の次の行（header_row_index が None の場合は 2）
                    start_row = header_row_index + 1 if header_row_index else 2
                    for row in ws.iter_rows(min_row=start_row, values_only=True):
                        if not row or len(row) <= code_col_index:
                            continue
                        row_code = str(row[code_col_index]).strip().upper() if row[code_col_index] else ""
                        if row_code == code:
                            result["found"] = True
                            matched_row = [convert_value(cell) for cell in row]
                            row_dict = {}
                            for i, header in enumerate(headers):
                                if i < len(matched_row):
                                    row_dict[header] = matched_row[i]
                                else:
                                    row_dict[header] = ""
                            matched_rows.append(row_dict)
                    result["data"][sheet_name] = {
                        "headers": headers,
                        "rows": matched_rows,
                        "count": len(matched_rows)
                    }
                    result["_meta"]["detection"][sheet_name] = detection_info
                    continue

                # 4) フォールバック：ヘッダー列が見つからない場合はシート全体をスキャンして
                #    任意のセルに code と一致する行を抽出する（scanned_by_value）
                detection_info['method'] = 'scanned_by_value'
                detection_info['header_row'] = header_row_index
                detection_info['code_col_index'] = None
                search_code_upper = code.upper().strip()
                for row in ws.iter_rows(min_row=(header_row_index + 1 if header_row_index else 2), values_only=True):
                    if not row:
                        continue
                    found = False
                    for cell in row:
                        if cell is None:
                            continue
                        try:
                            if str(cell).strip().upper() == search_code_upper:
                                found = True
                                break
                        except Exception:
                            continue
                    if found:
                        result["found"] = True
                        # row_dict を作成。ヘッダーがあればマッピング、無ければ col_x で返す
                        row_dict = {}
                        if headers:
                            for i, h in enumerate(headers):
                                if i < len(row):
                                    row_dict[h] = convert_value(row[i])
                                else:
                                    row_dict[h] = ""
                        else:
                            for i, val in enumerate(row):
                                row_dict[f"col_{i+1}"] = convert_value(val)
                        matched_rows.append(row_dict)

                result["data"][sheet_name] = {
                    "headers": headers,
                    "rows": matched_rows,
                    "count": len(matched_rows),
                    "note": "企業コード列が検出できなかったため、シート全体をスキャンして該当行を抽出しました（フォールバック）"
                }
                result["_meta"]["detection"][sheet_name] = detection_info

            # キャッシュして返却
            output_json = json.dumps(result, ensure_ascii=False, indent=2)
            set_cache(code, output_json)

            self.send_response(200)
            self.send_header('Content-Type', 'application/json; charset=utf-8')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            self.wfile.write(output_json.encode('utf-8'))

        except Exception as e:
            import traceback
            self.send_response(500)
            self.send_header('Content-Type', 'application/json; charset=utf-8')
            self.end_headers()
            error_response = json.dumps({
                "error": str(e),
                "type": type(e).__name__,
                "traceback": traceback.format_exc()
            }, ensure_ascii=False)
            self.wfile.write(error_response.encode('utf-8'))
