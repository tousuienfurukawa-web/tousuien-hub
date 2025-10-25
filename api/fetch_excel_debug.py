from http.server import BaseHTTPRequestHandler
from urllib.parse import urlparse, parse_qs
import json
from datetime import datetime, date

class DateTimeEncoder(json.JSONEncoder):
    """日付オブジェクトをJSON形式に変換するカスタムエンコーダー"""
    def default(self, obj):
        if isinstance(obj, (datetime, date)):
            return obj.isoformat()
        return super().default(obj)

def convert_value(value):
    """セルの値を安全な形式に変換"""
    if value is None:
        return ""
    elif isinstance(value, (datetime, date)):
        return value.isoformat()
    elif isinstance(value, (int, float, str, bool)):
        return value
    else:
        return str(value)

class handler(BaseHTTPRequestHandler):
    def do_GET(self):
        try:
            parsed_path = urlparse(self.path)
            params = parse_qs(parsed_path.query)
            
            if 'path' not in params:
                self.send_response(400)
                self.send_header('Content-Type', 'application/json; charset=utf-8')
                self.end_headers()
                self.wfile.write(json.dumps({"error": "Missing 'path' parameter"}).encode('utf-8'))
                return
            
            path = params['path'][0]
            
            import requests
            import io
            import openpyxl
            
            base_raw_url = "https://raw.githubusercontent.com/tousuienfurukawa-web/tousuien-hub/main/"
            url = base_raw_url + path
            
            response = requests.get(url, timeout=30)
            response.raise_for_status()
            
            workbook = openpyxl.load_workbook(io.BytesIO(response.content), data_only=True)
            result = {}
            
            # 各シートのヘッダーと最初の行を取得
            for sheet_name in workbook.sheetnames:
                ws = workbook[sheet_name]
                
                # ヘッダー行を取得（変換処理を適用）
                headers = []
                first_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))
                if first_row and first_row[0]:
                    headers = [convert_value(cell) for cell in first_row[0]]
                
                # データの最初の行を取得（変換処理を適用）
                first_data = []
                second_row = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))
                if second_row and second_row[0]:
                    first_data = [convert_value(cell) for cell in second_row[0]]
                
                result[sheet_name] = {
                    "headers": headers,
                    "first_row": first_data
                }
            
            self.send_response(200)
            self.send_header('Content-Type', 'application/json; charset=utf-8')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            
            # カスタムエンコーダーを使用
            output = json.dumps(result, ensure_ascii=False, indent=2, cls=DateTimeEncoder)
            self.wfile.write(output.encode('utf-8'))
            
        except Exception as e:
            import traceback
            self.send_response(500)
            self.send_header('Content-Type', 'application/json; charset=utf-8')
            self.end_headers()
            error_response = json.dumps({
                "error": str(e),
                "type": type(e).__name__,
                "traceback": traceback.format_exc()
            }, ensure_ascii=False)
            self.wfile.write(error_response.encode('utf-8'))
