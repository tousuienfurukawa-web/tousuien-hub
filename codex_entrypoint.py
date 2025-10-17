#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Codex-friendly entrypoint.

Usage (natural language):
  "TSE-BGV-058-25 出力して"
  "058-25 レポート"

It extracts the invoice id, uses default paths from environment variables
(or sensible fallbacks), then runs reconciliation and prints Markdown to stdout.

Environment variables:
- CM_EXCEL_PATH  (default: Customer Management_latest.xlsx)
- SLACK_ZIP_PATH (default: 海外 Slack export Feb 17 2022 - Oct 16 2025.zip)
- SLACK_DATE     (optional filter, e.g. 2025-09-24)
"""

import os
import re
import sys
from typing import Optional

# Local import without installing as a package
from scripts.reconcile_invoice import load_excel_record, scan_slack_zip, reconcile, as_markdown  # type: ignore


INVOICE_PATTERN = re.compile(r"\b([A-Z]{3}-[A-Z]{3}-\d{3}-\d{2})\b")
SUFFIX_PATTERN = re.compile(r"\b(\d{3}-\d{2})\b")


def infer_invoice_id(text: str, excel_path: str) -> Optional[str]:
    # Explicit invoice pattern first
    m = INVOICE_PATTERN.search(text)
    if m:
        return m.group(1)
    # If only suffix like 058-25 is provided, scan Excel to find a full invoice
    m2 = SUFFIX_PATTERN.search(text)
    if not m2:
        return None
    suffix = m2.group(1)
    try:
        from openpyxl import load_workbook  # type: ignore
        wb = load_workbook(excel_path, data_only=True, read_only=True)
        # try common orders sheet names
        candidates = ["受注登録", "受注", "受注台帳", "Orders", "Order", "Sales", "受注管理", "注文"]
        sheets = wb.sheetnames
        for s in candidates:
            if s in sheets:
                ws = wb[s]
                # detect header quickly (first 25 rows)
                header_row_idx = 1
                max_nonempty = -1
                for r in range(1, min(25, ws.max_row) + 1):
                    values = [c.value for c in ws[r]]
                    nonempty = sum(1 for v in values if ("" if v is None else str(v)).strip() != "")
                    if nonempty > max_nonempty:
                        max_nonempty = nonempty
                        header_row_idx = r
                headers = [("" if c.value is None else str(c.value)).strip() for c in ws[header_row_idx]]
                header_index = {h: i for i, h in enumerate(headers)}
                # find invoice column
                col_invoice = None
                for key in ("invoice", "インボイス", "請求書番号", "IV"):
                    if key in header_index:
                        col_invoice = header_index[key]
                        break
                if col_invoice is None:
                    continue
                for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
                    cells = ["" if v is None else str(v) for v in row]
                    inv = cells[col_invoice] if col_invoice < len(cells) else ""
                    if inv.endswith(suffix):
                        return inv
    except Exception:
        return None
    return None


def main() -> None:
    text = " ".join(sys.argv[1:])
    # Fallback to stdin if no args
    if not text.strip() and not sys.stdin.isatty():
        text = sys.stdin.read()

    excel_path = os.environ.get("CM_EXCEL_PATH", "Customer Management_latest.xlsx")
    slack_zip_path = os.environ.get("SLACK_ZIP_PATH", "海外 Slack export Feb 17 2022 - Oct 16 2025.zip")
    slack_date = os.environ.get("SLACK_DATE")

    invoice = infer_invoice_id(text, excel_path)
    if not invoice:
        # last resort: take the first token that looks like TSE-XXX-###-## or starts with TSE-
        m = re.search(r"\bTSE-[A-Z]{3}-\d{3}-\d{2}\b", text)
        invoice = m.group(0) if m else None

    if not invoice:
        print("請求書番号（例: TSE-BGV-058-25 または 058-25）が見つかりませんでした。", file=sys.stderr)
        sys.exit(2)

    excel = load_excel_record(excel_path, invoice)
    slack = scan_slack_zip(slack_zip_path, invoice, slack_date)
    result = reconcile(excel, slack)

    md = as_markdown(result)
    print(md)


if __name__ == "__main__":
    main()
