#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import argparse
import json
import os
import re
import sys
import zipfile
from typing import Dict, List, Optional, Any, Tuple

try:
    import openpyxl  # type: ignore
except Exception as exc:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl", file=sys.stderr)
    raise


def detect_orders_sheet(workbook: "openpyxl.Workbook") -> str:
    # Prefer exact then fuzzy Japanese terms
    preferred = [
        "受注登録",
        "受注",
        "受注台帳",
        "Orders",
        "Order",
        "Sales",
        "受注管理",
        "注文",
    ]
    sheets = workbook.sheetnames
    # Exact
    for name in sheets:
        if name in preferred:
            return name
    # Fuzzy
    for name in sheets:
        ln = name.lower()
        for kw in preferred:
            if kw.lower() in ln:
                return name
    # Fallback to the first sheet
    return sheets[0]


def detect_header_row(ws: "openpyxl.worksheet.worksheet.Worksheet", scan_rows: int = 25) -> Tuple[int, List[str]]:
    header_row_idx = 1
    max_nonempty = -1
    for r in range(1, min(scan_rows, ws.max_row) + 1):
        values = [c.value for c in ws[r]]
        nonempty = sum(1 for v in values if ("" if v is None else str(v)).strip() != "")
        if nonempty > max_nonempty:
            max_nonempty = nonempty
            header_row_idx = r
    headers = [ ("" if c.value is None else str(c.value)).strip() for c in ws[header_row_idx] ]
    return header_row_idx, headers


def col_index_map(headers: List[str]) -> Dict[str, int]:
    return {h: i for i, h in enumerate(headers)}


def find_col(index_by_name: Dict[str, int], *candidates: str) -> Optional[int]:
    # exact
    for nm in candidates:
        if nm in index_by_name:
            return index_by_name[nm]
    # fuzzy contains
    lower_index = {h.lower(): i for h, i in index_by_name.items()}
    for nm in candidates:
        lnm = nm.lower()
        for name_lower, idx in lower_index.items():
            if lnm in name_lower:
                return idx
    return None


def load_excel_record(excel_path: str, invoice_id: str) -> Dict[str, Any]:
    wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
    sheet_name = detect_orders_sheet(wb)
    ws = wb[sheet_name]

    header_row_idx, headers = detect_header_row(ws)
    index_by_name = col_index_map(headers)

    # Column heuristics
    col_invoice = find_col(index_by_name, "invoice", "インボイス", "請求書番号", "IV")
    col_company = find_col(index_by_name, "企業コード", "顧客コード", "会社コード")
    col_ship = find_col(index_by_name, "輸送方法", "発送方法", "配送方法", "Shipping")
    col_pay = find_col(index_by_name, "決済方法", "支払方法", "Payment")
    col_date = find_col(index_by_name, "注文日", "受注日", "日付")
    col_country = find_col(index_by_name, "輸出国", "国", "Country")
    col_bill_country = find_col(index_by_name, "請求国")
    col_addressee = find_col(index_by_name, "宛名")
    col_owner = find_col(index_by_name, "担当者名")

    def get(cells: List[str], idx: Optional[int]) -> str:
        return cells[idx] if (idx is not None and idx < len(cells)) else ""

    matched_rows: List[Dict[str, Any]] = []
    for ridx, row in enumerate(ws.iter_rows(min_row=header_row_idx + 1, values_only=True), start=header_row_idx + 1):
        cells = ["" if v is None else str(v) for v in row]
        inv = get(cells, col_invoice)
        if not inv:
            continue
        if invoice_id in inv:
            rec: Dict[str, Any] = {
                "row_index": ridx,
                "sheet": sheet_name,
                "invoice": inv,
                "company_code": get(cells, col_company),
                "shipping_method": get(cells, col_ship),
                "payment_method": get(cells, col_pay),
                "order_date": get(cells, col_date),
                "export_country": get(cells, col_country),
                "billing_country": get(cells, col_bill_country),
                "addressee": get(cells, col_addressee),
                "owner": get(cells, col_owner),
                "items": [],
            }
            # Collect product lines (up to 35)
            for i in range(1, 36):
                name_header = f"商品名{i}"
                pc_header = f"pc{i}"
                gpc_header = f"g/pc{i}"
                name_idx = index_by_name.get(name_header)
                pc_idx = index_by_name.get(pc_header)
                gpc_idx = index_by_name.get(gpc_header)
                name_val = get(cells, name_idx)
                pc_val = get(cells, pc_idx)
                gpc_val = get(cells, gpc_idx)
                if any([name_val, pc_val, gpc_val]):
                    rec["items"].append({
                        "name": name_val,
                        "pcs": pc_val,
                        "g_per_pc": gpc_val,
                    })
            matched_rows.append(rec)

    return {
        "excel_path": excel_path,
        "sheet": sheet_name,
        "headers": {i: h for i, h in enumerate(headers)},
        "invoice_id": invoice_id,
        "matches": matched_rows,
    }


def invoice_suffix(invoice_id: str) -> Optional[str]:
    parts = invoice_id.split("-")
    if len(parts) >= 2:
        return f"{parts[-2]}-{parts[-1]}"
    return None


def scan_slack_zip(zip_path: str, invoice_id: str, date: Optional[str] = None) -> Dict[str, Any]:
    suffix = invoice_suffix(invoice_id)
    terms = [invoice_id]
    if suffix and suffix not in terms:
        terms.append(suffix)
    # Add supportive terms often used in reconciliation
    supportive_terms = [
        "PayPal", "DHL", "FedEx", "受注", "入金", "支払い", "請求", "ラベル", "AWB",
    ]
    terms += supportive_terms

    results: List[Dict[str, Any]] = []
    files_scanned = 0

    with zipfile.ZipFile(zip_path, "r") as zf:
        names = [n for n in zf.namelist() if n.endswith(".json")]
        if date:
            names = [n for n in names if n.endswith(f"{date}.json")]
        for name in names:
            try:
                content = zf.read(name).decode("utf-8")
            except UnicodeDecodeError:
                content = zf.read(name).decode("utf-8", errors="replace")
            files_scanned += 1
            try:
                arr = json.loads(content)
            except Exception:
                continue
            hits: List[Dict[str, Any]] = []
            for m in arr:
                txt: str = (m.get("text") or "")
                if any(t in txt for t in terms):
                    hits.append({
                        "path": name,
                        "ts": m.get("ts"),
                        "user": m.get("user"),
                        "thread_ts": m.get("thread_ts"),
                        "reply_count": m.get("reply_count"),
                        "text": txt,
                        "user_profile": (m.get("user_profile") or {}),
                    })
            if hits:
                results.extend(hits)

    # Group by thread_ts if present
    threads: Dict[str, List[Dict[str, Any]]] = {}
    for msg in results:
        tts = msg.get("thread_ts") or msg.get("ts")
        if not tts:
            continue
        threads.setdefault(tts, []).append(msg)

    # Extract simple signals from messages
    shipping_mentions = set()
    payment_mentions = set()
    amounts_usd: List[str] = []

    for msg in results:
        t = msg.get("text", "")
        if "DHL" in t:
            shipping_mentions.add("DHL")
        if re.search(r"\bFedEx\b", t, re.IGNORECASE):
            shipping_mentions.add("FedEx")
        if "PayPal" in t:
            payment_mentions.add("PayPal")
        # capture like: 387.90 USD or 4,539.75 USD
        for m in re.finditer(r"([0-9][0-9,]*\.?[0-9]*)\s*USD", t):
            amounts_usd.append(m.group(1) + " USD")

    return {
        "zip_path": zip_path,
        "date_filter": date,
        "invoice_id": invoice_id,
        "files_scanned": files_scanned,
        "message_hits": results,
        "threads": threads,
        "shipping_mentions": sorted(shipping_mentions),
        "payment_mentions": sorted(payment_mentions),
        "amounts_usd": amounts_usd,
    }


def reconcile(excel: Dict[str, Any], slack: Dict[str, Any]) -> Dict[str, Any]:
    excel_match = excel.get("matches", [{}])[0] if excel.get("matches") else {}

    # Basic comparisons
    excel_shipping = (excel_match.get("shipping_method") or "").strip()
    excel_payment = (excel_match.get("payment_method") or "").strip()
    excel_date = (excel_match.get("order_date") or "").strip()
    excel_company = (excel_match.get("company_code") or "").strip()

    # Signals from Slack
    slack_shipping = slack.get("shipping_mentions", [])
    slack_payments = slack.get("payment_mentions", [])

    checks = []

    # Shipping
    if excel_shipping:
        checks.append({
            "field": "shipping_method",
            "excel": excel_shipping,
            "slack_signal": ", ".join(slack_shipping) if slack_shipping else None,
            "match": (excel_shipping in slack_shipping) if slack_shipping else None,
        })
    else:
        checks.append({
            "field": "shipping_method",
            "excel": None,
            "slack_signal": ", ".join(slack_shipping) if slack_shipping else None,
            "match": None,
            "note": "Excel empty; Slack suggests: " + ", ".join(slack_shipping) if slack_shipping else None,
        })

    # Payment
    if excel_payment:
        checks.append({
            "field": "payment_method",
            "excel": excel_payment,
            "slack_signal": ", ".join(slack_payments) if slack_payments else None,
            "match": (excel_payment in slack_payments) if slack_payments else None,
        })
    else:
        checks.append({
            "field": "payment_method",
            "excel": None,
            "slack_signal": ", ".join(slack_payments) if slack_payments else None,
            "match": None,
            "note": "Excel empty; Slack suggests: " + ", ".join(slack_payments) if slack_payments else None,
        })

    # Summarize items
    items = excel_match.get("items") or []
    summary = {
        "invoice": excel.get("invoice_id"),
        "company_code": excel_company or None,
        "order_date": excel_date or None,
        "shipping_method_excel": excel_shipping or None,
        "payment_method_excel": excel_payment or None,
        "items_count": len(items),
        "items": items,
        "slack_shipping_mentions": slack_shipping,
        "slack_payment_mentions": slack_payments,
        "slack_usd_amounts": slack.get("amounts_usd", []),
        "checks": checks,
    }

    # Simple overall status
    status = "ok"
    for c in checks:
        if c["match"] is False:
            status = "mismatch"
            break
    if not checks:
        status = "unknown"

    return {
        "status": status,
        "excel": excel,
        "slack": slack,
        "summary": summary,
    }


def as_markdown(result: Dict[str, Any]) -> str:
    s = result.get("summary", {})
    lines = []
    lines.append(f"# Reconciliation for {s.get('invoice')}")
    lines.append("")
    lines.append(f"- Company code: {s.get('company_code')}")
    lines.append(f"- Order date: {s.get('order_date')}")
    lines.append(f"- Excel shipping: {s.get('shipping_method_excel')}")
    lines.append(f"- Excel payment: {s.get('payment_method_excel')}")
    if s.get("slack_shipping_mentions"):
        lines.append(f"- Slack shipping mentions: {', '.join(s['slack_shipping_mentions'])}")
    if s.get("slack_payment_mentions"):
        lines.append(f"- Slack payment mentions: {', '.join(s['slack_payment_mentions'])}")
    if s.get("slack_usd_amounts"):
        lines.append(f"- Slack USD amounts: {', '.join(s['slack_usd_amounts'])}")
    lines.append("")
    lines.append("## Items")
    if s.get("items"):
        for it in s["items"]:
            name = it.get("name")
            pcs = it.get("pcs")
            gpc = it.get("g_per_pc")
            lines.append(f"- {name} × {pcs} (g/pc: {gpc})")
    else:
        lines.append("- (no items parsed)")
    lines.append("")
    lines.append("## Checks")
    for c in s.get("checks", []):
        field = c.get("field")
        match = c.get("match")
        excel = c.get("excel")
        slack_signal = c.get("slack_signal")
        note = c.get("note")
        if match is True:
            state = "MATCH"
        elif match is False:
            state = "MISMATCH"
        else:
            state = "UNKNOWN"
        line = f"- {field}: {state} (excel: {excel}, slack: {slack_signal})"
        if note:
            line += f" — {note}"
        lines.append(line)
    return "\n".join(lines)


def main() -> None:
    parser = argparse.ArgumentParser(description="Reconcile Excel '受注登録' vs Slack export for a given invoice.")
    parser.add_argument("--excel", required=True, help="Path to Customer Management Excel file")
    parser.add_argument("--slack-zip", required=True, help="Path to Slack export ZIP")
    parser.add_argument("--invoice", required=True, help="Invoice ID, e.g., TSE-BGV-058-25")
    parser.add_argument("--slack-date", default=None, help="Optional Slack date filter YYYY-MM-DD (narrows files)")
    parser.add_argument("--format", choices=["json", "md"], default="json", help="Output format")
    parser.add_argument("--out-file", default=None, help="Optional path to write output; default stdout")

    args = parser.parse_args()

    excel_data = load_excel_record(args.excel, args.invoice)
    slack_data = scan_slack_zip(args.slack_zip, args.invoice, args.slack_date)
    result = reconcile(excel_data, slack_data)

    if args.format == "json":
        text = json.dumps(result, ensure_ascii=False, indent=2)
    else:
        text = as_markdown(result)

    if args.out_file:
        os.makedirs(os.path.dirname(args.out_file) or ".", exist_ok=True)
        with open(args.out_file, "w", encoding="utf-8") as f:
            f.write(text)
    else:
        print(text)


if __name__ == "__main__":
    main()
